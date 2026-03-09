import openpyxl
import json
import os

# 加载配置文件
config_path = os.path.join(os.path.dirname(__file__), 'config.json')
with open(config_path, 'r', encoding='utf-8') as f:
    config = json.load(f)

# 读取 Excel 文件
wb = openpyxl.load_workbook(
    config['waterfall_excel_path'],
    data_only=True
)
ws = wb.active
# 读取 U2 和 V2 单元格的值
u2_value = ws['U2'].value  # U列是上月
v2_value = ws['V2'].value  # V列是当月

# 转换日期格式为"年月"
def format_date(date_value):
    from datetime import datetime, timedelta

    # 处理 Excel 日期序列号（45992 这种）
    if isinstance(date_value, (int, float)):
        base_date = datetime(1899, 12, 30)
        dt = base_date + timedelta(days=int(date_value))
        return f"{dt.year}年{dt.month}月"

    # 处理字符串
    if isinstance(date_value, str):
        try:
            dt = datetime.strptime(date_value, '%Y/%m/%d')
            return f"{dt.year}年{dt.month}月"
        except:
            return str(date_value)

    # 处理 datetime 类型
    if hasattr(date_value, 'year') and hasattr(date_value, 'month'):
        return f"{date_value.year}年{date_value.month}月"

    return str(date_value)

u2_formatted = format_date(u2_value)  # 上月
v2_formatted = format_date(v2_value)  # 当月

# 读取台次环比变化数据 (B8:B13, U8:U13, V8:V13)
taici_categories = []
taici_current = []
taici_last = []

for row in range(8, 14):  # 8到13行
    cat = ws[f'B{row}'].value
    last = ws[f'U{row}'].value  # U列是上月
    current = ws[f'V{row}'].value  # V列是当月

    if cat:
        taici_categories.append(str(cat))
        taici_current.append(float(current) if current is not None else 0)
        taici_last.append(float(last) if last is not None else 0)

# 读取混合维修环比变化数据 (B22:B27, U22:U27, V22:V27)
hunhe_categories = []
hunhe_current = []
hunhe_last = []

for row in range(22, 28):  # 22到27行
    cat = ws[f'B{row}'].value
    last = ws[f'U{row}'].value  # U列是上月
    current = ws[f'V{row}'].value  # V列是当月

    if cat:
        hunhe_categories.append(str(cat))
        hunhe_current.append(float(current) if current is not None else 0)
        hunhe_last.append(float(last) if last is not None else 0)

# 读取内部维修环比变化数据 (C37:C44, U37:U44, V37:V44)
neibu_categories = []
neibu_current = []
neibu_last = []

for row in range(37, 45):  # 37到44行
    cat = ws[f'C{row}'].value
    last = ws[f'U{row}'].value  # U列是上月
    current = ws[f'V{row}'].value  # V列是当月

    if cat:
        neibu_categories.append(str(cat))
        neibu_current.append(float(current) if current is not None else 0)
        neibu_last.append(float(last) if last is not None else 0)

# 读取门店运营成本分析数据 (C45:C137, U45:U137, V45:V137)
mendian_data = []

for row in range(45, 138):  # 45到137行
    cat = ws[f'C{row}'].value
    last = ws[f'U{row}'].value  # U列是上月
    current = ws[f'V{row}'].value  # V列是当月

    # 获取单元格样式信息来判断层级
    cell = ws[f'C{row}']
    is_bold = cell.font.bold if cell.font else False
    fill_color = cell.fill.fgColor.rgb if cell.fill and hasattr(cell.fill.fgColor, 'rgb') else None

    # 跳过错误值和无效数据
    if current is None and last is None:
        continue
    if str(current) == '#REF!' or str(last) == '#REF!':
        continue

    # 判断层级
    level = 0
    if not cat and is_bold:  # 汇总行（无类别名称但加粗）
        level = 0
        cat = "合计"
    elif is_bold and fill_color == 'FFDEE0E3':  # 一级分类（加粗+灰色）
        level = 1
    elif fill_color == 'FFFAF1D1':  # 二级分类（浅黄色）
        level = 2
    else:  # 三级分类
        level = 3

    if cat:  # 只添加有类别名称的行
        try:
            mendian_data.append({
                'category': str(cat),
                'current': float(current) if current is not None and str(current) != '#REF!' else 0,
                'last': float(last) if last is not None and str(last) != '#REF!' else 0,
                'level': level
            })
        except (ValueError, TypeError):
            # 跳过无法转换的值
            continue

# 读取高频高毛利TOP20项目数据
from collections import defaultdict

wb2 = openpyxl.load_workbook(config['detail_excel_path'], data_only=True)
ws2 = wb2.active

# 统计所有账期
period_set = set()
for i in range(2, ws2.max_row + 1):
    a_val = ws2.cell(row=i, column=1).value
    if a_val:
        period_set.add(str(a_val))

# 按账期排序，取最新的两个
periods = sorted(list(period_set), reverse=True)
current_period = periods[0] if len(periods) > 0 else ""
last_period = periods[1] if len(periods) > 1 else ""

# 第一步：收集所有"轮胎"相关的维修项目编码
tire_project_codes = set()
for i in range(2, ws2.max_row + 1):
    ae_val = ws2.cell(row=i, column=31).value  # AE列 - 商品类型名称
    ad_val = ws2.cell(row=i, column=30).value  # AD列 - 商品名称
    x_val = ws2.cell(row=i, column=24).value   # X列 - 维修项目编码

    # 如果商品类型=零件 且 商品名称=轮胎，记录维修项目编码
    if str(ae_val) == "零件" and str(ad_val) == "轮胎" and x_val:
        tire_project_codes.add(str(x_val))

# 按项目统计
project_stats = defaultdict(lambda: {'current_qty': 0, 'last_qty': 0, 'current_maoli': 0, 'last_maoli': 0})

for i in range(2, ws2.max_row + 1):
    a_val = str(ws2.cell(row=i, column=1).value) if ws2.cell(row=i, column=1).value else ""
    x_val = ws2.cell(row=i, column=24).value  # X列 - 维修项目编码
    y_val = ws2.cell(row=i, column=25).value  # Y列 - 项目名称
    af_val = ws2.cell(row=i, column=32).value  # AF列 - 收入类型
    ag_val = ws2.cell(row=i, column=33).value  # AG列 - 商品数量
    ak_val = ws2.cell(row=i, column=37).value  # AK列 - 实收
    aw_val = ws2.cell(row=i, column=49).value  # AW列 - 主机厂零件实收

    # 过滤条件：项目名称存在 且 收入类型不等于"混合维修"
    if not y_val or str(af_val) == "混合维修":
        continue

    # 特殊处理：如果维修项目编码在轮胎项目编码集合中，统一使用"轮胎"作为项目名称
    if x_val and str(x_val) in tire_project_codes:
        project_name = "轮胎"
    else:
        project_name = str(y_val)

    qty = float(ag_val) if ag_val else 0
    maoli = (float(ak_val) if ak_val else 0) - (float(aw_val) if aw_val else 0)

    if a_val == current_period:
        project_stats[project_name]['current_qty'] += qty
        project_stats[project_name]['current_maoli'] += maoli
    elif a_val == last_period:
        project_stats[project_name]['last_qty'] += qty
        project_stats[project_name]['last_maoli'] += maoli

# 计算总毛利并排序，取TOP20
top20_projects = []
tire_project = None  # 保存轮胎项目

for project, stats in project_stats.items():
    total_maoli = stats['current_maoli'] + stats['last_maoli']
    if abs(total_maoli) > 0.01:  # 只看有毛利的项目
        project_data = {
            'name': project,
            'current_qty': stats['current_qty'],
            'last_qty': stats['last_qty'],
            'current_maoli': stats['current_maoli'],
            'last_maoli': stats['last_maoli'],
            'total_maoli': total_maoli
        }
        top20_projects.append(project_data)

        # 如果是轮胎项目，单独保存
        if project == "轮胎":
            tire_project = project_data

top20_projects.sort(key=lambda x: x['total_maoli'], reverse=True)
top20_projects = top20_projects[:20]  # 取TOP20

# 强制确保轮胎项目在榜单中
if tire_project and tire_project not in top20_projects:
    # 如果轮胎不在TOP20中，强制加入
    top20_projects.append(tire_project)
    print(f"注意：轮胎项目（毛利: {tire_project['total_maoli']:.2f}）强制加入榜单")

# 统计混合维修的TOP20项目（按商品名称）
hunhe_product_stats = defaultdict(lambda: {'current_qty': 0, 'last_qty': 0, 'current_maoli': 0, 'last_maoli': 0})

for i in range(2, ws2.max_row + 1):
    a_val = str(ws2.cell(row=i, column=1).value) if ws2.cell(row=i, column=1).value else ""
    ad_val = ws2.cell(row=i, column=30).value  # AD列 - 商品名称
    af_val = ws2.cell(row=i, column=32).value  # AF列 - 收入类型
    ag_val = ws2.cell(row=i, column=33).value  # AG列 - 商品数量
    ak_val = ws2.cell(row=i, column=37).value  # AK列 - 实收
    aw_val = ws2.cell(row=i, column=49).value  # AW列 - 主机厂零件实收

    # 过滤条件：商品名称存在 且 收入类型等于"混合维修"
    if not ad_val or str(af_val) != "混合维修":
        continue

    qty = float(ag_val) if ag_val else 0
    maoli = (float(ak_val) if ak_val else 0) - (float(aw_val) if aw_val else 0)

    if a_val == current_period:
        hunhe_product_stats[ad_val]['current_qty'] += qty
        hunhe_product_stats[ad_val]['current_maoli'] += maoli
    elif a_val == last_period:
        hunhe_product_stats[ad_val]['last_qty'] += qty
        hunhe_product_stats[ad_val]['last_maoli'] += maoli

# 计算总毛利并排序，取TOP20
top20_hunhe_products = []
for product, stats in hunhe_product_stats.items():
    total_maoli = stats['current_maoli'] + stats['last_maoli']
    if abs(total_maoli) > 0.01:
        top20_hunhe_products.append({
            'name': product,
            'current_qty': stats['current_qty'],
            'last_qty': stats['last_qty'],
            'current_maoli': stats['current_maoli'],
            'last_maoli': stats['last_maoli'],
            'total_maoli': total_maoli
        })

top20_hunhe_products.sort(key=lambda x: x['total_maoli'], reverse=True)
top20_hunhe_products = top20_hunhe_products[:20]  # 取TOP20

# 统计内部维修的TOP10项目（按收入类型分类）
# 4.2.1 保修-质保TOP10
baozheng_zhbao_stats = defaultdict(lambda: {'current_qty': 0, 'last_qty': 0, 'current_maoli': 0, 'last_maoli': 0})

for i in range(2, ws2.max_row + 1):
    a_val = str(ws2.cell(row=i, column=1).value) if ws2.cell(row=i, column=1).value else ""
    y_val = ws2.cell(row=i, column=25).value  # Y列 - 项目名称
    af_val = ws2.cell(row=i, column=32).value  # AF列 - 收入类型
    ag_val = ws2.cell(row=i, column=33).value  # AG列 - 商品数量
    an_val = ws2.cell(row=i, column=40).value  # AN列 - 内部结算收入

    if not y_val or str(af_val) != "保修-质保":
        continue

    qty = float(ag_val) if ag_val else 0
    maoli = float(an_val) if an_val else 0

    if a_val == current_period:
        baozheng_zhbao_stats[y_val]['current_qty'] += qty
        baozheng_zhbao_stats[y_val]['current_maoli'] += maoli
    elif a_val == last_period:
        baozheng_zhbao_stats[y_val]['last_qty'] += qty
        baozheng_zhbao_stats[y_val]['last_maoli'] += maoli

top10_baozheng_zhbao = []
for project, stats in baozheng_zhbao_stats.items():
    total_maoli = stats['current_maoli'] + stats['last_maoli']
    if abs(total_maoli) > 0.01:
        top10_baozheng_zhbao.append({
            'name': project,
            'current_qty': stats['current_qty'],
            'last_qty': stats['last_qty'],
            'current_maoli': stats['current_maoli'],
            'last_maoli': stats['last_maoli'],
            'total_maoli': total_maoli
        })
top10_baozheng_zhbao.sort(key=lambda x: x['total_maoli'], reverse=True)
top10_baozheng_zhbao = top10_baozheng_zhbao[:10]

# 4.2.2 保修-技术升级TOP10
jishu_shengji_stats = defaultdict(lambda: {'current_qty': 0, 'last_qty': 0, 'current_maoli': 0, 'last_maoli': 0})

for i in range(2, ws2.max_row + 1):
    a_val = str(ws2.cell(row=i, column=1).value) if ws2.cell(row=i, column=1).value else ""
    y_val = ws2.cell(row=i, column=25).value
    af_val = ws2.cell(row=i, column=32).value
    ag_val = ws2.cell(row=i, column=33).value
    an_val = ws2.cell(row=i, column=40).value  # AN列 - 内部结算收入

    if not y_val or str(af_val) != "保修-技术升级":
        continue

    qty = float(ag_val) if ag_val else 0
    maoli = float(an_val) if an_val else 0

    if a_val == current_period:
        jishu_shengji_stats[y_val]['current_qty'] += qty
        jishu_shengji_stats[y_val]['current_maoli'] += maoli
    elif a_val == last_period:
        jishu_shengji_stats[y_val]['last_qty'] += qty
        jishu_shengji_stats[y_val]['last_maoli'] += maoli

top10_jishu_shengji = []
for project, stats in jishu_shengji_stats.items():
    total_maoli = stats['current_maoli'] + stats['last_maoli']
    if abs(total_maoli) > 0.01:
        top10_jishu_shengji.append({
            'name': project,
            'current_qty': stats['current_qty'],
            'last_qty': stats['last_qty'],
            'current_maoli': stats['current_maoli'],
            'last_maoli': stats['last_maoli'],
            'total_maoli': total_maoli
        })
top10_jishu_shengji.sort(key=lambda x: x['total_maoli'], reverse=True)
top10_jishu_shengji = top10_jishu_shengji[:10]

# 4.2.3 保修-终身质保TOP10
zhongshen_zhbao_stats = defaultdict(lambda: {'current_qty': 0, 'last_qty': 0, 'current_maoli': 0, 'last_maoli': 0})

for i in range(2, ws2.max_row + 1):
    a_val = str(ws2.cell(row=i, column=1).value) if ws2.cell(row=i, column=1).value else ""
    y_val = ws2.cell(row=i, column=25).value
    af_val = ws2.cell(row=i, column=32).value
    ag_val = ws2.cell(row=i, column=33).value
    an_val = ws2.cell(row=i, column=40).value  # AN列 - 内部结算收入

    if not y_val or str(af_val) != "保修-终身质保":
        continue

    qty = float(ag_val) if ag_val else 0
    maoli = float(an_val) if an_val else 0

    if a_val == current_period:
        zhongshen_zhbao_stats[y_val]['current_qty'] += qty
        zhongshen_zhbao_stats[y_val]['current_maoli'] += maoli
    elif a_val == last_period:
        zhongshen_zhbao_stats[y_val]['last_qty'] += qty
        zhongshen_zhbao_stats[y_val]['last_maoli'] += maoli

top10_zhongshen_zhbao = []
for project, stats in zhongshen_zhbao_stats.items():
    total_maoli = stats['current_maoli'] + stats['last_maoli']
    if abs(total_maoli) > 0.01:
        top10_zhongshen_zhbao.append({
            'name': project,
            'current_qty': stats['current_qty'],
            'last_qty': stats['last_qty'],
            'current_maoli': stats['current_maoli'],
            'last_maoli': stats['last_maoli'],
            'total_maoli': total_maoli
        })
top10_zhongshen_zhbao.sort(key=lambda x: x['total_maoli'], reverse=True)
top10_zhongshen_zhbao = top10_zhongshen_zhbao[:10]

# 4.2.4 服务产品TOP10
fuwu_chanpin_stats = defaultdict(lambda: {'current_qty': 0, 'last_qty': 0, 'current_maoli': 0, 'last_maoli': 0})

for i in range(2, ws2.max_row + 1):
    a_val = str(ws2.cell(row=i, column=1).value) if ws2.cell(row=i, column=1).value else ""
    y_val = ws2.cell(row=i, column=25).value
    af_val = ws2.cell(row=i, column=32).value
    ag_val = ws2.cell(row=i, column=33).value
    an_val = ws2.cell(row=i, column=40).value  # AN列 - 内部结算收入

    if not y_val or str(af_val) != "服务产品":
        continue

    qty = float(ag_val) if ag_val else 0
    maoli = float(an_val) if an_val else 0

    if a_val == current_period:
        fuwu_chanpin_stats[y_val]['current_qty'] += qty
        fuwu_chanpin_stats[y_val]['current_maoli'] += maoli
    elif a_val == last_period:
        fuwu_chanpin_stats[y_val]['last_qty'] += qty
        fuwu_chanpin_stats[y_val]['last_maoli'] += maoli

top10_fuwu_chanpin = []
for project, stats in fuwu_chanpin_stats.items():
    total_maoli = stats['current_maoli'] + stats['last_maoli']
    if abs(total_maoli) > 0.01:
        top10_fuwu_chanpin.append({
            'name': project,
            'current_qty': stats['current_qty'],
            'last_qty': stats['last_qty'],
            'current_maoli': stats['current_maoli'],
            'last_maoli': stats['last_maoli'],
            'total_maoli': total_maoli
        })
top10_fuwu_chanpin.sort(key=lambda x: x['total_maoli'], reverse=True)
top10_fuwu_chanpin = top10_fuwu_chanpin[:10]

# 4.2.5 商城安装TOP10（按商品名称AD列）
shangcheng_anzhuang_stats = defaultdict(lambda: {'current_qty': 0, 'last_qty': 0, 'current_maoli': 0, 'last_maoli': 0})

for i in range(2, ws2.max_row + 1):
    a_val = str(ws2.cell(row=i, column=1).value) if ws2.cell(row=i, column=1).value else ""
    ad_val = ws2.cell(row=i, column=30).value  # AD列 - 商品名称
    af_val = ws2.cell(row=i, column=32).value
    ag_val = ws2.cell(row=i, column=33).value
    an_val = ws2.cell(row=i, column=40).value  # AN列 - 内部结算收入

    if not ad_val or str(af_val) != "商城安装":
        continue

    qty = float(ag_val) if ag_val else 0
    maoli = float(an_val) if an_val else 0

    if a_val == current_period:
        shangcheng_anzhuang_stats[ad_val]['current_qty'] += qty
        shangcheng_anzhuang_stats[ad_val]['current_maoli'] += maoli
    elif a_val == last_period:
        shangcheng_anzhuang_stats[ad_val]['last_qty'] += qty
        shangcheng_anzhuang_stats[ad_val]['last_maoli'] += maoli

top10_shangcheng_anzhuang = []
for product, stats in shangcheng_anzhuang_stats.items():
    total_maoli = stats['current_maoli'] + stats['last_maoli']
    if abs(total_maoli) > 0.01:
        top10_shangcheng_anzhuang.append({
            'name': product,
            'current_qty': stats['current_qty'],
            'last_qty': stats['last_qty'],
            'current_maoli': stats['current_maoli'],
            'last_maoli': stats['last_maoli'],
            'total_maoli': total_maoli
        })
top10_shangcheng_anzhuang.sort(key=lambda x: x['total_maoli'], reverse=True)
top10_shangcheng_anzhuang = top10_shangcheng_anzhuang[:10]

# 读取 AE:AG 列的数据（从第8行开始）
categories = []
values = []
labels = []

for i, row in enumerate(ws.iter_rows(min_col=31, max_col=33, values_only=True), start=1):
    # 跳过前7行（标题和空行）
    if i < 8:
        continue
    # 如果类别和数值都有，就添加
    if row[0] is not None and row[1] is not None:
        categories.append(str(row[0]))
        values.append(float(row[1]))
        labels.append(str(row[2]) if row[2] is not None else '')
    # 如果遇到空行就停止
    elif row[0] is None and row[1] is None and row[2] is None:
        break

# 计算瀑布图数据
cumulative = []
current_sum = 0

for i, val in enumerate(values):
    if i == 0:
        cumulative.append(0)
        current_sum = val
    elif i == len(values) - 1:  # 最后一个是总计
        cumulative.append(0)
    else:
        cumulative.append(current_sum)
        current_sum += val

# 分类分析数据（使用 AG 列的标签）
analysis_data = {
    '收入端-维修保养': [],
    '收入端-混合维修': [],
    '收入端-内部结算': [],
    '费用端': []
}

for cat, val, label in zip(categories, values, labels):
    if label in analysis_data:
        analysis_data[label].append({'name': cat, 'value': val})

# 计算总结数据
total_weibao = sum(item['value'] for item in analysis_data['收入端-维修保养'])
total_hunhe = sum(item['value'] for item in analysis_data['收入端-混合维修'])
total_neibu = sum(item['value'] for item in analysis_data['收入端-内部结算'])
total_feiyong = sum(item['value'] for item in analysis_data['费用端'])

total_revenue = total_weibao + total_hunhe + total_neibu
net_profit = values[-1] if len(values) > 0 else 0  # 最后一个值是总利润

# 判断表现描述
def get_performance(value):
    if value > 50000:
        return "表现优秀"
    elif value > 0:
        return "表现良好"
    elif value > -50000:
        return "略有下降"
    else:
        return "显著下降"


# 生成 HTML
html_content = f"""<!DOCTYPE html>
<html>
<head>
    <meta charset="utf-8">
    <title>瀑布图综合分析</title>
    <script src="https://cdn.jsdelivr.net/npm/echarts@5.4.3/dist/echarts.min.js"></script>
    <style>
        body {{
            margin: 0;
            padding: 20px;
            font-family: 'Microsoft YaHei', Arial, sans-serif;
            background-color: #f5f5f5;
        }}
        .container {{
            max-width: 1400px;
            margin: 0 auto;
            background-color: white;
            padding: 30px;
            border-radius: 8px;
            box-shadow: 0 2px 8px rgba(0,0,0,0.1);
        }}
        h1 {{
            text-align: center;
            color: #333;
            margin-bottom: 30px;
        }}
        .section-title {{
            color: #333;
            font-size: 20px;
            font-weight: bold;
            margin-top: 30px;
            margin-bottom: 15px;
            padding-left: 10px;
            border-left: 5px solid #4CAF50;
        }}
        h2 {{
            color: #555;
            margin-top: 30px;
            margin-bottom: 15px;
            border-bottom: 2px solid #4CAF50;
            padding-bottom: 8px;
        }}
        .chart {{
            width: 100%;
            height: 600px;
            margin-bottom: 40px;
        }}
        .analysis-grid {{
            display: grid;
            grid-template-columns: repeat(2, 1fr);
            gap: 20px;
            margin-top: 25px;
        }}
        .analysis-item {{
            background: #fafafa;
            padding: 15px;
            border-radius: 8px;
            border: 1px solid #e0e0e0;
        }}
        .analysis-item h3 {{
            margin-top: 0;
            color: #333;
            font-size: 16px;
            border-left: 4px solid #4CAF50;
            padding-left: 10px;
        }}
        .stat {{
            display: flex;
            justify-content: space-between;
            padding: 6px 0;
            border-bottom: 1px solid #e0e0e0;
        }}
        .stat:last-child {{
            border-bottom: none;
        }}
        .stat-label {{
            color: #666;
            font-size: 13px;
        }}
        .stat-value {{
            font-weight: bold;
            color: #333;
            font-size: 13px;
        }}
        .stat-value.positive {{
            color: #4CAF50;
        }}
        .stat-value.negative {{
            color: #f44336;
        }}
        .summary-section {{
            background: linear-gradient(135deg, #e0e7ff 0%, #f3e8ff 100%);
            color: #1f2937;
            padding: 20px;
            border-radius: 8px;
            margin-bottom: 25px;
            border: 1px solid #c7d2fe;
        }}
        .summary-title {{
            font-size: 18px;
            font-weight: bold;
            margin-bottom: 12px;
            border-bottom: 2px solid #a5b4fc;
            padding-bottom: 8px;
            color: #4338ca;
        }}
        .summary-item {{
            background: rgba(255,255,255,0.6);
            padding: 10px 12px;
            border-radius: 6px;
            margin-bottom: 10px;
            border-left: 3px solid #8b5cf6;
        }}
        .summary-item h4 {{
            margin: 0 0 6px 0;
            font-size: 14px;
            color: #4338ca;
            font-weight: 600;
        }}
        .summary-item p {{
            margin: 3px 0;
            font-size: 13px;
            line-height: 1.5;
            color: #374151;
        }}
        .summary-number {{
            font-size: 15px;
            font-weight: bold;
        }}
        .summary-number.positive {{
            color: #16a34a;
        }}
        .summary-number.negative {{
            color: #dc2626;
        }}
        .conclusion-box {{
            background: rgba(255,255,255,0.7);
            padding: 12px 15px;
            border-radius: 6px;
            margin-top: 10px;
            border: 2px solid #a78bfa;
        }}
        .conclusion-box h3 {{
            margin: 0 0 6px 0;
            font-size: 15px;
            color: #7c3aed;
            font-weight: 600;
        }}
        .conclusion-box p {{
            margin: 0;
            font-size: 13px;
            line-height: 1.6;
            color: #374151;
        }}
        .data-table {{
            width: 100%;
            border-collapse: collapse;
            margin-top: 15px;
            background: white;
            border-radius: 8px;
            overflow: hidden;
            box-shadow: 0 1px 3px rgba(0,0,0,0.1);
        }}
        .data-table th {{
            background: rgb(0, 114, 109);
            color: white;
            padding: 12px 8px;
            text-align: center;
            font-size: 14px;
            font-weight: 600;
        }}
        .data-table td {{
            padding: 10px 8px;
            text-align: center;
            border-bottom: 1px solid #e0e0e0;
            font-size: 13px;
        }}
        .data-table tr:last-child td {{
            border-bottom: none;
        }}
        .data-table tr:hover {{
            background: #f5f5f5;
        }}
        .data-table td.category {{
            text-align: left;
            font-weight: 500;
        }}
        .data-table td.positive {{
            color: #16a34a;
            font-weight: bold;
        }}
        .data-table td.negative {{
            color: #dc2626;
            font-weight: bold;
        }}
        .top10-grid {{
            display: grid;
            grid-template-columns: 1fr 1fr;
            gap: 20px;
            margin-top: 20px;
        }}
        .top10-item {{
            min-width: 0;
            display: flex;
            flex-direction: column;
        }}
        .top10-item .data-table {{
            flex: 1;
            height: 100%;
        }}
        .top10-item .data-table tbody {{
            min-height: 400px;
        }}
        @media (max-width: 1200px) {{
            .top10-grid {{
                grid-template-columns: 1fr;
            }}
        }}
        .mendian-table tr[style*='background: #EFEFEF'] {{
            font-weight: bold;
            background: #EFEFEF !important;
        }}
        .mendian-table tr[style*='background: #FFF9E6'] {{
            background: #FFF9E6 !important;
        }}
        .mendian-table .category[style*='padding-left: 20px'] {{
            position: relative;
        }}
        .mendian-table .category[style*='padding-left: 20px']::before {{
            content: '└─';
            color: #999;
            margin-right: 5px;
        }}
        .mendian-table .category[style*='padding-left: 40px'] {{
            position: relative;
        }}
        .mendian-table .category[style*='padding-left: 40px']::before {{
            content: '　└─';
            color: #ccc;
            margin-right: 5px;
        }}
    </style>
</head>
<body>
    <div class="container">
        <h1>📊 瀑布图综合分析报告</h1>

        <!-- 总结和核心结论 -->
        <div class="summary-section">
            <div class="summary-title">环比洞察总结</div>

            <div class="summary-item">
                <h4>📈 整体趋势</h4>
                <p>整体{('向好' if net_profit > 0 else '承压')}，利润环比变化 <span class="summary-number {('positive' if net_profit > 0 else 'negative')}">{net_profit:+,.2f}元</span></p>
            </div>

            <div class="summary-item">
                <h4>💰 收入端表现</h4>
                <p>{('正向贡献' if total_revenue > 0 else '负向影响')} <span class="summary-number {('positive' if total_revenue > 0 else 'negative')}">{total_revenue:+,.2f}元</span></p>
                <p>• 维保业务：<span class="summary-number {('positive' if total_weibao > 0 else 'negative')}">{total_weibao:+,.2f}元</span>（{get_performance(total_weibao)}）</p>
                <p>• 混合维修：<span class="summary-number {('positive' if total_hunhe > 0 else 'negative')}">{total_hunhe:+,.2f}元</span>（{get_performance(total_hunhe)}）</p>
                <p>• 内部结算：<span class="summary-number {('positive' if total_neibu > 0 else 'negative')}">{total_neibu:+,.2f}元</span>（{get_performance(total_neibu)}）</p>
            </div>

            <div class="summary-item">
                <h4>💸 费用端表现</h4>
                <p>{('正向贡献' if total_feiyong > 0 else '负向影响')} <span class="summary-number {('positive' if total_feiyong > 0 else 'negative')}">{total_feiyong:+,.2f}元</span></p>
                <p>费用控制{'得当' if total_feiyong > 0 else '需要改善'}，对利润形成{('正向支持' if total_feiyong > 0 else '负向压力')}</p>
            </div>

            <div class="conclusion-box">
                <h3>🎯 核心结论</h3>
                <p>费用端贡献（<span class="summary-number {('positive' if total_feiyong > 0 else 'negative')}">{total_feiyong:+,.2f}元</span>）{'有效弥补了' if total_feiyong > 0 and total_revenue < 0 else '叠加'}收入端{'压力' if total_revenue < 0 else '增长'}（<span class="summary-number {('positive' if total_revenue > 0 else 'negative')}">{total_revenue:+,.2f}元</span>），最终实现净{'增长' if net_profit > 0 else '下降'}<span class="summary-number {('positive' if net_profit > 0 else 'negative')}">{net_profit:+,.2f}元</span>，整体经营{('向好' if net_profit > 0 else '需要关注')}。</p>
            </div>
        </div>

        <div class="section-title">一、利润环比瀑布图分析({v2_formatted} VS {u2_formatted})</div>

        <h2>整体瀑布图</h2>
        <div id="waterfall-chart" class="chart"></div>

        <div class="section-title">1.1、分类详细分析</div>

        <div class="analysis-grid">
            <div class="analysis-item">
                <h3 style="border-left-color: rgb(0, 114, 109);">收入端-维修保养</h3>
                <div id="weibao-stats"></div>
                <div id="weibao-chart" style="width: 100%; height: 240px;"></div>
            </div>

            <div class="analysis-item">
                <h3 style="border-left-color: rgb(206, 164, 114);">收入端-混合维修</h3>
                <div id="hunhe-stats"></div>
                <div id="hunhe-chart" style="width: 100%; height: 240px;"></div>
            </div>

            <div class="analysis-item">
                <h3 style="border-left-color: rgb(0, 114, 109);">收入端-内部结算</h3>
                <div id="neibu-stats"></div>
                <div id="neibu-chart" style="width: 100%; height: 240px;"></div>
            </div>

            <div class="analysis-item">
                <h3 style="border-left-color: rgb(244, 67, 54);">费用端</h3>
                <div id="feiyong-stats"></div>
                <div id="feiyong-chart" style="width: 100%; height: 240px;"></div>
            </div>
        </div>

        <div class="section-title">二、维修保养专项分析</div>

        <div class="section-title" style="border-left-color: rgb(0, 114, 109);">2.1 环比变化</div>
        <table class="data-table">
            <thead>
                <tr>
                    <th>类别</th>
                    <th>当月</th>
                    <th>上月</th>
                    <th>变化</th>
                    <th>变化率</th>
                </tr>
            </thead>
            <tbody>"""

# 生成表格行
for i, cat in enumerate(taici_categories):
    current = taici_current[i]
    last = taici_last[i]
    change = current - last
    change_rate = (change / last * 100) if last != 0 else 0
    change_class = 'positive' if change > 0 else 'negative' if change < 0 else ''

    html_content += f"""
                <tr>
                    <td class="category">{cat}</td>
                    <td>{current:.2f}</td>
                    <td>{last:.2f}</td>
                    <td class="{change_class}">{change:+.2f}</td>
                    <td class="{change_class}">{change_rate:+.2f}%</td>
                </tr>"""

html_content += f"""
            </tbody>
        </table>

        <div class="section-title" style="border-left-color: rgb(0, 114, 109); margin-top: 40px;">2.2 高频高毛利TOP20项目环比对比</div>
        <div style="background: #fff3cd; border-left: 4px solid #ffc107; padding: 12px 15px; margin-bottom: 15px; border-radius: 4px;">
            <span style="font-size: 18px;">💡</span>
            <span style="color: #856404; font-size: 14px; margin-left: 8px;">
                <strong>分析说明：</strong>以下展示基于总毛利排序的TOP20维保项目的环比对比，帮助识别维保业务增长或下降的主要原因，当月/上月均取值于工时零件商品数量而非台次，下同
            </span>
        </div>
        <table class="data-table">
            <thead>
                <tr>
                    <th style="width: 50px;">排名</th>
                    <th style="width: 180px;">项目名称</th>
                    <th>当月</th>
                    <th>上月</th>
                    <th>变化</th>
                    <th>变化率</th>
                    <th>当月毛利</th>
                    <th>上月毛利</th>
                    <th>毛利变化</th>
                    <th>变化率</th>
                </tr>
            </thead>
            <tbody>"""

# 生成TOP20项目表格行
for rank, proj in enumerate(top20_projects, 1):
    qty_change = proj['current_qty'] - proj['last_qty']
    qty_rate = (qty_change / proj['last_qty'] * 100) if proj['last_qty'] != 0 else 0
    maoli_change = proj['current_maoli'] - proj['last_maoli']
    maoli_rate = (maoli_change / proj['last_maoli'] * 100) if proj['last_maoli'] != 0 else 0

    qty_change_class = 'positive' if qty_change > 0 else 'negative' if qty_change < 0 else ''
    maoli_change_class = 'positive' if maoli_change > 0 else 'negative' if maoli_change < 0 else ''

    html_content += f"""
                <tr>
                    <td style="text-align: center; font-weight: bold; color: #666;">{rank}</td>
                    <td class="category">{proj['name']}</td>
                    <td>{proj['current_qty']:.2f}</td>
                    <td>{proj['last_qty']:.2f}</td>
                    <td class="{qty_change_class}">{qty_change:+.2f}</td>
                    <td class="{qty_change_class}">{qty_rate:+.1f}%</td>
                    <td>{proj['current_maoli']:.2f}</td>
                    <td>{proj['last_maoli']:.2f}</td>
                    <td class="{maoli_change_class}">{maoli_change:+.2f}</td>
                    <td class="{maoli_change_class}">{maoli_rate:+.1f}%</td>
                </tr>"""

html_content += f"""
            </tbody>
        </table>

        <div class="section-title" style="margin-top: 50px;">三、混合维修专项分析</div>

        <div class="section-title" style="border-left-color: rgb(0, 114, 109);">3.1 环比变化</div>
        <table class="data-table">
            <thead>
                <tr>
                    <th>类别</th>
                    <th>当月</th>
                    <th>上月</th>
                    <th>变化</th>
                    <th>变化率</th>
                </tr>
            </thead>
            <tbody>"""

# 生成混合维修表格行
for i, cat in enumerate(hunhe_categories):
    current = hunhe_current[i]
    last = hunhe_last[i]
    change = current - last
    change_rate = (change / last * 100) if last != 0 else 0
    change_class = 'positive' if change > 0 else 'negative' if change < 0 else ''

    html_content += f"""
                <tr>
                    <td class="category">{cat}</td>
                    <td>{current:.2f}</td>
                    <td>{last:.2f}</td>
                    <td class="{change_class}">{change:+.2f}</td>
                    <td class="{change_class}">{change_rate:+.2f}%</td>
                </tr>"""

html_content += f"""
            </tbody>
        </table>

        <div class="section-title" style="border-left-color: rgb(0, 114, 109); margin-top: 40px;">3.2 高频高毛利TOP20项目环比对比</div>
        <div style="background: #fff3cd; border-left: 4px solid #ffc107; padding: 12px 15px; margin-bottom: 15px; border-radius: 4px;">
            <span style="font-size: 18px;">💡</span>
            <span style="color: #856404; font-size: 14px; margin-left: 8px;">
                <strong>分析说明：</strong>以下展示基于总毛利排序的TOP20混合维修商品的环比对比，帮助识别混合维修业务增长或下降的主要原因
            </span>
        </div>
        <table class="data-table">
            <thead>
                <tr>
                    <th style="width: 50px;">排名</th>
                    <th style="width: 180px;">商品名称</th>
                    <th>当月</th>
                    <th>上月</th>
                    <th>变化</th>
                    <th>变化率</th>
                    <th>当月毛利</th>
                    <th>上月毛利</th>
                    <th>毛利变化</th>
                    <th>变化率</th>
                </tr>
            </thead>
            <tbody>"""

# 生成混合维修TOP20商品表格行
for rank, prod in enumerate(top20_hunhe_products, 1):
    qty_change = prod['current_qty'] - prod['last_qty']
    qty_rate = (qty_change / prod['last_qty'] * 100) if prod['last_qty'] != 0 else 0
    maoli_change = prod['current_maoli'] - prod['last_maoli']
    maoli_rate = (maoli_change / prod['last_maoli'] * 100) if prod['last_maoli'] != 0 else 0

    qty_change_class = 'positive' if qty_change > 0 else 'negative' if qty_change < 0 else ''
    maoli_change_class = 'positive' if maoli_change > 0 else 'negative' if maoli_change < 0 else ''

    html_content += f"""
                <tr>
                    <td style="text-align: center; font-weight: bold; color: #666;">{rank}</td>
                    <td class="category">{prod['name']}</td>
                    <td>{prod['current_qty']:.2f}</td>
                    <td>{prod['last_qty']:.2f}</td>
                    <td class="{qty_change_class}">{qty_change:+.2f}</td>
                    <td class="{qty_change_class}">{qty_rate:+.1f}%</td>
                    <td>{prod['current_maoli']:.2f}</td>
                    <td>{prod['last_maoli']:.2f}</td>
                    <td class="{maoli_change_class}">{maoli_change:+.2f}</td>
                    <td class="{maoli_change_class}">{maoli_rate:+.1f}%</td>
                </tr>"""

html_content += f"""
            </tbody>
        </table>

        <div class="section-title" style="margin-top: 50px;">四、内部维修专项分析</div>

        <div class="section-title" style="border-left-color: rgb(0, 114, 109);">4.1 环比变化</div>
        <table class="data-table">
            <thead>
                <tr>
                    <th>类别</th>
                    <th>当月</th>
                    <th>上月</th>
                    <th>变化</th>
                    <th>变化率</th>
                </tr>
            </thead>
            <tbody>"""

# 生成内部维修表格行
for i, cat in enumerate(neibu_categories):
    current = neibu_current[i]
    last = neibu_last[i]
    change = current - last
    change_rate = (change / last * 100) if last != 0 else 0
    change_class = 'positive' if change > 0 else 'negative' if change < 0 else ''

    html_content += f"""
                <tr>
                    <td class="category">{cat}</td>
                    <td>{current:.2f}</td>
                    <td>{last:.2f}</td>
                    <td class="{change_class}">{change:+.2f}</td>
                    <td class="{change_class}">{change_rate:+.2f}%</td>
                </tr>"""

html_content += f"""
            </tbody>
        </table>

        <div class="section-title" style="border-left-color: rgb(0, 114, 109); margin-top: 40px;">4.2 TOP10项目环比对比</div>

        <div class="top10-grid">
            <div class="top10-item">
                <div class="section-title" style="border-left-color: rgb(0, 114, 109); margin-top: 0; font-size: 16px;">4.2.1 保修-质保TOP10</div>
        <table class="data-table">
            <thead>
                <tr>
                    <th style="width: 50px;">排名</th>
                    <th style="width: 180px;">项目名称</th>
                    <th>当月</th>
                    <th>上月</th>
                    <th>变化</th>
                    <th>变化率</th>
                    <th>当月毛利</th>
                    <th>上月毛利</th>
                    <th>毛利变化</th>
                    <th>变化率</th>
                </tr>
            </thead>
            <tbody>"""

# 生成保修-质保TOP10表格行
for rank in range(1, 11):  # 固定显示10行
    if rank <= len(top10_baozheng_zhbao):
        proj = top10_baozheng_zhbao[rank - 1]
        qty_change = proj['current_qty'] - proj['last_qty']
        qty_rate = (qty_change / proj['last_qty'] * 100) if proj['last_qty'] != 0 else 0
        maoli_change = proj['current_maoli'] - proj['last_maoli']
        maoli_rate = (maoli_change / proj['last_maoli'] * 100) if proj['last_maoli'] != 0 else 0

        qty_change_class = 'positive' if qty_change > 0 else 'negative' if qty_change < 0 else ''
        maoli_change_class = 'positive' if maoli_change > 0 else 'negative' if maoli_change < 0 else ''

        html_content += f"""
                <tr>
                    <td style="text-align: center; font-weight: bold; color: #666;">{rank}</td>
                    <td class="category">{proj['name']}</td>
                    <td>{proj['current_qty']:.2f}</td>
                    <td>{proj['last_qty']:.2f}</td>
                    <td class="{qty_change_class}">{qty_change:+.2f}</td>
                    <td class="{qty_change_class}">{qty_rate:+.1f}%</td>
                    <td>{proj['current_maoli']:.2f}</td>
                    <td>{proj['last_maoli']:.2f}</td>
                    <td class="{maoli_change_class}">{maoli_change:+.2f}</td>
                    <td class="{maoli_change_class}">{maoli_rate:+.1f}%</td>
                </tr>"""
    else:
        # 填充空行
        html_content += f"""
                <tr>
                    <td style="text-align: center; font-weight: bold; color: #666;">{rank}</td>
                    <td class="category" style="color: #ccc;">-</td>
                    <td style="color: #ccc;">-</td>
                    <td style="color: #ccc;">-</td>
                    <td style="color: #ccc;">-</td>
                    <td style="color: #ccc;">-</td>
                    <td style="color: #ccc;">-</td>
                    <td style="color: #ccc;">-</td>
                    <td style="color: #ccc;">-</td>
                    <td style="color: #ccc;">-</td>
                </tr>"""

html_content += f"""
            </tbody>
        </table>
            </div>

            <div class="top10-item">
                <div class="section-title" style="border-left-color: rgb(0, 114, 109); margin-top: 0; font-size: 16px;">4.2.2 保修-技术升级TOP10</div>
        <table class="data-table">
            <thead>
                <tr>
                    <th style="width: 50px;">排名</th>
                    <th style="width: 180px;">项目名称</th>
                    <th>当月</th>
                    <th>上月</th>
                    <th>变化</th>
                    <th>变化率</th>
                    <th>当月毛利</th>
                    <th>上月毛利</th>
                    <th>毛利变化</th>
                    <th>变化率</th>
                </tr>
            </thead>
            <tbody>"""

# 生成保修-技术升级TOP10表格行
for rank in range(1, 11):  # 固定显示10行
    if rank <= len(top10_jishu_shengji):
        proj = top10_jishu_shengji[rank - 1]
        qty_change = proj['current_qty'] - proj['last_qty']
        qty_rate = (qty_change / proj['last_qty'] * 100) if proj['last_qty'] != 0 else 0
        maoli_change = proj['current_maoli'] - proj['last_maoli']
        maoli_rate = (maoli_change / proj['last_maoli'] * 100) if proj['last_maoli'] != 0 else 0

        qty_change_class = 'positive' if qty_change > 0 else 'negative' if qty_change < 0 else ''
        maoli_change_class = 'positive' if maoli_change > 0 else 'negative' if maoli_change < 0 else ''

        html_content += f"""
                <tr>
                    <td style="text-align: center; font-weight: bold; color: #666;">{rank}</td>
                    <td class="category">{proj['name']}</td>
                    <td>{proj['current_qty']:.2f}</td>
                    <td>{proj['last_qty']:.2f}</td>
                    <td class="{qty_change_class}">{qty_change:+.2f}</td>
                    <td class="{qty_change_class}">{qty_rate:+.1f}%</td>
                    <td>{proj['current_maoli']:.2f}</td>
                    <td>{proj['last_maoli']:.2f}</td>
                    <td class="{maoli_change_class}">{maoli_change:+.2f}</td>
                    <td class="{maoli_change_class}">{maoli_rate:+.1f}%</td>
                </tr>"""
    else:
        # 填充空行
        html_content += f"""
                <tr>
                    <td style="text-align: center; font-weight: bold; color: #666;">{rank}</td>
                    <td class="category" style="color: #ccc;">-</td>
                    <td style="color: #ccc;">-</td>
                    <td style="color: #ccc;">-</td>
                    <td style="color: #ccc;">-</td>
                    <td style="color: #ccc;">-</td>
                    <td style="color: #ccc;">-</td>
                    <td style="color: #ccc;">-</td>
                    <td style="color: #ccc;">-</td>
                    <td style="color: #ccc;">-</td>
                </tr>"""

html_content += f"""
            </tbody>
        </table>
            </div>

            <div class="top10-item">
                <div class="section-title" style="border-left-color: rgb(0, 114, 109); margin-top: 0; font-size: 16px;">4.2.3 保修-终身质保TOP10</div>
        <table class="data-table">
            <thead>
                <tr>
                    <th style="width: 50px;">排名</th>
                    <th style="width: 180px;">项目名称</th>
                    <th>当月</th>
                    <th>上月</th>
                    <th>变化</th>
                    <th>变化率</th>
                    <th>当月毛利</th>
                    <th>上月毛利</th>
                    <th>毛利变化</th>
                    <th>变化率</th>
                </tr>
            </thead>
            <tbody>"""

# 生成保修-终身质保TOP10表格行
for rank in range(1, 11):  # 固定显示10行
    if rank <= len(top10_zhongshen_zhbao):
        proj = top10_zhongshen_zhbao[rank - 1]
        qty_change = proj['current_qty'] - proj['last_qty']
        qty_rate = (qty_change / proj['last_qty'] * 100) if proj['last_qty'] != 0 else 0
        maoli_change = proj['current_maoli'] - proj['last_maoli']
        maoli_rate = (maoli_change / proj['last_maoli'] * 100) if proj['last_maoli'] != 0 else 0

        qty_change_class = 'positive' if qty_change > 0 else 'negative' if qty_change < 0 else ''
        maoli_change_class = 'positive' if maoli_change > 0 else 'negative' if maoli_change < 0 else ''

        html_content += f"""
                <tr>
                    <td style="text-align: center; font-weight: bold; color: #666;">{rank}</td>
                    <td class="category">{proj['name']}</td>
                    <td>{proj['current_qty']:.2f}</td>
                    <td>{proj['last_qty']:.2f}</td>
                    <td class="{qty_change_class}">{qty_change:+.2f}</td>
                    <td class="{qty_change_class}">{qty_rate:+.1f}%</td>
                    <td>{proj['current_maoli']:.2f}</td>
                    <td>{proj['last_maoli']:.2f}</td>
                    <td class="{maoli_change_class}">{maoli_change:+.2f}</td>
                    <td class="{maoli_change_class}">{maoli_rate:+.1f}%</td>
                </tr>"""
    else:
        # 填充空行
        html_content += f"""
                <tr>
                    <td style="text-align: center; font-weight: bold; color: #666;">{rank}</td>
                    <td class="category" style="color: #ccc;">-</td>
                    <td style="color: #ccc;">-</td>
                    <td style="color: #ccc;">-</td>
                    <td style="color: #ccc;">-</td>
                    <td style="color: #ccc;">-</td>
                    <td style="color: #ccc;">-</td>
                    <td style="color: #ccc;">-</td>
                    <td style="color: #ccc;">-</td>
                    <td style="color: #ccc;">-</td>
                </tr>"""

html_content += f"""
            </tbody>
        </table>
            </div>

            <div class="top10-item">
                <div class="section-title" style="border-left-color: rgb(0, 114, 109); margin-top: 0; font-size: 16px;">4.2.4 服务产品TOP10</div>
        <table class="data-table">
            <thead>
                <tr>
                    <th style="width: 50px;">排名</th>
                    <th style="width: 180px;">项目名称</th>
                    <th>当月</th>
                    <th>上月</th>
                    <th>变化</th>
                    <th>变化率</th>
                    <th>当月毛利</th>
                    <th>上月毛利</th>
                    <th>毛利变化</th>
                    <th>变化率</th>
                </tr>
            </thead>
            <tbody>"""

# 生成服务产品TOP10表格行
for rank in range(1, 11):  # 固定显示10行
    if rank <= len(top10_fuwu_chanpin):
        proj = top10_fuwu_chanpin[rank - 1]
        qty_change = proj['current_qty'] - proj['last_qty']
        qty_rate = (qty_change / proj['last_qty'] * 100) if proj['last_qty'] != 0 else 0
        maoli_change = proj['current_maoli'] - proj['last_maoli']
        maoli_rate = (maoli_change / proj['last_maoli'] * 100) if proj['last_maoli'] != 0 else 0

        qty_change_class = 'positive' if qty_change > 0 else 'negative' if qty_change < 0 else ''
        maoli_change_class = 'positive' if maoli_change > 0 else 'negative' if maoli_change < 0 else ''

        html_content += f"""
                <tr>
                    <td style="text-align: center; font-weight: bold; color: #666;">{rank}</td>
                    <td class="category">{proj['name']}</td>
                    <td>{proj['current_qty']:.2f}</td>
                    <td>{proj['last_qty']:.2f}</td>
                    <td class="{qty_change_class}">{qty_change:+.2f}</td>
                    <td class="{qty_change_class}">{qty_rate:+.1f}%</td>
                    <td>{proj['current_maoli']:.2f}</td>
                    <td>{proj['last_maoli']:.2f}</td>
                    <td class="{maoli_change_class}">{maoli_change:+.2f}</td>
                    <td class="{maoli_change_class}">{maoli_rate:+.1f}%</td>
                </tr>"""
    else:
        # 填充空行
        html_content += f"""
                <tr>
                    <td style="text-align: center; font-weight: bold; color: #666;">{rank}</td>
                    <td class="category" style="color: #ccc;">-</td>
                    <td style="color: #ccc;">-</td>
                    <td style="color: #ccc;">-</td>
                    <td style="color: #ccc;">-</td>
                    <td style="color: #ccc;">-</td>
                    <td style="color: #ccc;">-</td>
                    <td style="color: #ccc;">-</td>
                    <td style="color: #ccc;">-</td>
                    <td style="color: #ccc;">-</td>
                </tr>"""

html_content += f"""
            </tbody>
        </table>
            </div>
        </div>

        <div class="section-title" style="border-left-color: rgb(0, 114, 109); margin-top: 30px; font-size: 16px;">4.2.5 商城安装TOP10</div>
        <table class="data-table">
            <thead>
                <tr>
                    <th style="width: 50px;">排名</th>
                    <th style="width: 180px;">商品名称</th>
                    <th>当月</th>
                    <th>上月</th>
                    <th>变化</th>
                    <th>变化率</th>
                    <th>当月毛利</th>
                    <th>上月毛利</th>
                    <th>毛利变化</th>
                    <th>变化率</th>
                </tr>
            </thead>
            <tbody>"""

# 生成商城安装TOP10表格行
for rank in range(1, 11):  # 固定显示10行
    if rank <= len(top10_shangcheng_anzhuang):
        prod = top10_shangcheng_anzhuang[rank - 1]
        qty_change = prod['current_qty'] - prod['last_qty']
        qty_rate = (qty_change / prod['last_qty'] * 100) if prod['last_qty'] != 0 else 0
        maoli_change = prod['current_maoli'] - prod['last_maoli']
        maoli_rate = (maoli_change / prod['last_maoli'] * 100) if prod['last_maoli'] != 0 else 0

        qty_change_class = 'positive' if qty_change > 0 else 'negative' if qty_change < 0 else ''
        maoli_change_class = 'positive' if maoli_change > 0 else 'negative' if maoli_change < 0 else ''

        html_content += f"""
                <tr>
                    <td style="text-align: center; font-weight: bold; color: #666;">{rank}</td>
                    <td class="category">{prod['name']}</td>
                    <td>{prod['current_qty']:.2f}</td>
                    <td>{prod['last_qty']:.2f}</td>
                    <td class="{qty_change_class}">{qty_change:+.2f}</td>
                    <td class="{qty_change_class}">{qty_rate:+.1f}%</td>
                    <td>{prod['current_maoli']:.2f}</td>
                    <td>{prod['last_maoli']:.2f}</td>
                    <td class="{maoli_change_class}">{maoli_change:+.2f}</td>
                    <td class="{maoli_change_class}">{maoli_rate:+.1f}%</td>
                </tr>"""
    else:
        # 填充空行
        html_content += f"""
                <tr>
                    <td style="text-align: center; font-weight: bold; color: #666;">{rank}</td>
                    <td class="category" style="color: #ccc;">-</td>
                    <td style="color: #ccc;">-</td>
                    <td style="color: #ccc;">-</td>
                    <td style="color: #ccc;">-</td>
                    <td style="color: #ccc;">-</td>
                    <td style="color: #ccc;">-</td>
                    <td style="color: #ccc;">-</td>
                    <td style="color: #ccc;">-</td>
                    <td style="color: #ccc;">-</td>
                </tr>"""

html_content += f"""
            </tbody>
        </table>

        <div class="section-title" style="margin-top: 50px;">五、门店运营成本分析</div>"""

# 生成门店运营成本分析结论（层级分析法）
# 按层级组织数据
level1_costs = {}
level2_costs = {}
level3_costs = {}

for item in mendian_data:
    change = item['current'] - item['last']
    change_rate = (change / item['last'] * 100) if item['last'] != 0 else 0

    cost_data = {
        'category': item['category'],
        'current': item['current'],
        'last': item['last'],
        'change': change,
        'change_rate': change_rate
    }

    if item['level'] == 1:
        level1_costs[item['category']] = cost_data
        level1_costs[item['category']]['children'] = []
    elif item['level'] == 2:
        level2_costs[item['category']] = cost_data
        level2_costs[item['category']]['children'] = []
    elif item['level'] == 3:
        level3_costs[item['category']] = cost_data

# 建立层级关系（简单方式：按顺序关联）
current_level1 = None
current_level2 = None
for item in mendian_data:
    if item['level'] == 1:
        current_level1 = item['category']
        current_level2 = None
    elif item['level'] == 2 and current_level1:
        current_level2 = item['category']
        if current_level1 in level1_costs and item['category'] in level2_costs:
            level1_costs[current_level1]['children'].append(level2_costs[item['category']])
    elif item['level'] == 3 and current_level2:
        if current_level2 in level2_costs and item['category'] in level3_costs:
            level2_costs[current_level2]['children'].append(level3_costs[item['category']])

# 计算总成本变化
total_current = sum(c['current'] for c in level1_costs.values())
total_last = sum(c['last'] for c in level1_costs.values())
total_change = total_current - total_last
total_change_rate = (total_change / total_last * 100) if total_last != 0 else 0

# 按变化金额排序一级分类
level1_sorted = sorted(level1_costs.values(), key=lambda x: abs(x['change']), reverse=True)

# 生成层级分析结论
conclusion_html = f"<p>本月门店运营成本总额为<strong>{total_current:,.0f}元</strong>，较上月"
if total_change > 0:
    conclusion_html += f"<strong>增加{abs(total_change):,.0f}元</strong>（<strong style='color: #d9534f;'>+{total_change_rate:.1f}%</strong>）。"
elif total_change < 0:
    conclusion_html += f"<strong>减少{abs(total_change):,.0f}元</strong>（<strong style='color: #5cb85c;'>{total_change_rate:.1f}%</strong>）。"
else:
    conclusion_html += "基本持平。"

conclusion_html += "主要变化来自：</p><ul style='margin: 8px 0; padding-left: 20px;'>"

# 展示前2个变化最大的一级分类及其细节
for idx, level1_item in enumerate(level1_sorted[:2]):
    if abs(level1_item['change']) < 100:  # 跳过变化太小的
        continue

    # 一级分类
    color_style = 'color: #d9534f;' if level1_item['change'] > 0 else 'color: #5cb85c;'
    change_text = f"增加{abs(level1_item['change']):,.0f}元" if level1_item['change'] > 0 else f"减少{abs(level1_item['change']):,.0f}元"

    conclusion_html += f"<li style='margin-bottom: 8px;'><strong>{level1_item['category']}</strong>"
    conclusion_html += f"{change_text}（<span style='{color_style}'>{level1_item['change_rate']:+.1f}%</span>）"

    # 展示该一级分类下变化较大的二级/三级明细
    if level1_item['children']:
        # 找出变化最大的2个子项
        children_sorted = sorted(level1_item['children'], key=lambda x: abs(x['change']), reverse=True)
        detail_items = []
        for child in children_sorted[:2]:
            if abs(child['change']) > 50:  # 只显示变化较大的
                child_text = f"{child['category']}"
                if child['change'] > 0:
                    child_text += f"增加{abs(child['change']):,.0f}元"
                else:
                    child_text += f"减少{abs(child['change']):,.0f}元"
                detail_items.append(child_text)

        if detail_items:
            conclusion_html += "，其中" + "、".join(detail_items)

    conclusion_html += "</li>"

conclusion_html += "</ul>"

mendian_conclusion = conclusion_html

html_content += f"""
        <div style="background: #e8f5e9; border-left: 4px solid #4caf50; padding: 15px 20px; margin: 20px 0; border-radius: 4px; line-height: 1.8;">
            <div style="display: flex; align-items: flex-start;">
                <span style="font-size: 24px; margin-right: 10px;">💰</span>
                <div style="flex: 1;">
                    <div style="font-weight: bold; color: #2e7d32; margin-bottom: 8px; font-size: 15px;">成本分析结论</div>
                    <div style="color: #1b5e20; font-size: 14px;">{mendian_conclusion}</div>
                </div>
            </div>
        </div>
        <table class="data-table mendian-table">
            <thead>
                <tr>
                    <th style="width: 350px;">类别</th>
                    <th style="width: 120px;">当月</th>
                    <th style="width: 120px;">上月</th>
                    <th style="width: 120px;">变化</th>
                    <th style="width: 120px;">变化率</th>
                </tr>
            </thead>
            <tbody>"""

# 生成门店运营成本表格行
for item in mendian_data:
    current = item['current']
    last = item['last']
    change = current - last
    change_rate = (change / last * 100) if last != 0 else 0
    # 成本的颜色逻辑：增加显示红色，减少显示绿色
    change_class = 'negative' if change > 0 else 'positive' if change < 0 else ''

    # 根据层级设置样式
    if item['level'] == 1:  # 一级分类（加粗+灰色背景）
        row_style = 'font-weight: bold; background: #EFEFEF;'
        cat_style = 'font-weight: bold;'
    elif item['level'] == 2:  # 二级分类（浅黄背景+左缩进）
        row_style = 'background: #FFF9E6;'
        cat_style = 'padding-left: 20px; font-weight: 500;'
    else:  # 三级分类（白色背景+更多缩进）
        row_style = ''
        cat_style = 'padding-left: 40px;'

    html_content += f"""
                <tr style="{row_style}">
                    <td class="category" style="{cat_style}">{item['category']}</td>
                    <td>{current:.2f}</td>
                    <td>{last:.2f}</td>
                    <td class="{change_class}">{change:+.2f}</td>
                    <td class="{change_class}">{change_rate:+.2f}%</td>
                </tr>"""

html_content += f"""
            </tbody>
        </table>
    </div>

    <script>
        // 瀑布图数据
        var categories = {json.dumps(categories, ensure_ascii=False)};
        var values = {json.dumps(values)};
        var cumulative = {json.dumps(cumulative)};

        // 分析数据
        var analysisData = {json.dumps(analysis_data, ensure_ascii=False)};

        // 绘制主瀑布图
        var waterfallChart = echarts.init(document.getElementById('waterfall-chart'));
        var invisible = [];
        var data = [];

        for (var i = 0; i < values.length; i++) {{
            invisible.push(cumulative[i]);
            data.push(values[i]);
        }}

        var waterfallOption = {{
            tooltip: {{
                trigger: 'axis',
                axisPointer: {{
                    type: 'shadow'
                }},
                formatter: function(params) {{
                    var value = params[1].value;
                    var cumValue = invisible[params[0].dataIndex] + value;
                    return params[0].name + '<br/>' +
                           '变化: ' + value.toFixed(2) + '<br/>' +
                           '累计: ' + cumValue.toFixed(2);
                }}
            }},
            grid: {{
                left: '3%',
                right: '4%',
                bottom: '15%',
                containLabel: true
            }},
            xAxis: {{
                type: 'category',
                data: categories,
                axisLabel: {{
                    interval: 0,
                    rotate: 45,
                    fontSize: 11
                }}
            }},
            yAxis: {{
                type: 'value',
                axisLabel: {{
                    formatter: '{{value}}'
                }}
            }},
            series: [
                {{
                    name: '辅助',
                    type: 'bar',
                    stack: 'total',
                    itemStyle: {{
                        borderColor: 'transparent',
                        color: 'transparent'
                    }},
                    emphasis: {{
                        itemStyle: {{
                            borderColor: 'transparent',
                            color: 'transparent'
                        }}
                    }},
                    data: invisible
                }},
                {{
                    name: '数值',
                    type: 'bar',
                    stack: 'total',
                    label: {{
                        show: true,
                        position: 'inside',
                        formatter: function(params) {{
                            return params.value.toFixed(0);
                        }},
                        fontSize: 10
                    }},
                    itemStyle: {{
                        color: function(params) {{
                            if (params.dataIndex === values.length - 1) {{
                                return '#5470c6';
                            }}
                            return params.value > 0 ? '#91cc75' : '#ee6666';
                        }}
                    }},
                    data: data
                }}
            ]
        }};

        waterfallChart.setOption(waterfallOption);

        // 绘制分析图表的函数
        function drawAnalysisChart(elementId, data, title, baseColor) {{
            if (data.length === 0) {{
                document.getElementById(elementId).innerHTML = '<div style="text-align: center; padding: 50px; color: #999;">暂无数据</div>';
                return;
            }}

            var chart = echarts.init(document.getElementById(elementId));

            // 根据基础颜色生成渐变色系
            var colors = generateColorPalette(baseColor, data.length);

            var option = {{
                tooltip: {{
                    trigger: 'item',
                    formatter: function(params) {{
                        return params.name + ': ' + params.value.toFixed(2) + ' (' + params.percent.toFixed(2) + '%)';
                    }}
                }},
                color: colors,
                series: [
                    {{
                        type: 'pie',
                        radius: ['40%', '70%'],
                        avoidLabelOverlap: false,
                        itemStyle: {{
                            borderRadius: 10,
                            borderColor: '#fff',
                            borderWidth: 2
                        }},
                        label: {{
                            show: true,
                            formatter: function(params) {{
                                return params.name + '\\n' + params.value.toFixed(2);
                            }}
                        }},
                        data: data.map(item => ({{
                            name: item.name,
                            value: Math.abs(item.value)
                        }}))
                    }}
                ]
            }};
            chart.setOption(option);
        }}

        // 生成颜色渐变系列
        function generateColorPalette(baseColor, count) {{
            var colors = [baseColor];
            if (count > 1) {{
                // 解析 RGB
                var rgb = baseColor.match(/\\d+/g).map(Number);
                for (var i = 1; i < count; i++) {{
                    var factor = 0.3 + (i / count) * 0.5;
                    var r = Math.min(255, Math.floor(rgb[0] + (255 - rgb[0]) * factor));
                    var g = Math.min(255, Math.floor(rgb[1] + (255 - rgb[1]) * factor));
                    var b = Math.min(255, Math.floor(rgb[2] + (255 - rgb[2]) * factor));
                    colors.push('rgb(' + r + ',' + g + ',' + b + ')');
                }}
            }}
            return colors;
        }}

        // 显示统计信息的函数
        function showStats(elementId, data, color) {{
            if (data.length === 0) {{
                document.getElementById(elementId).innerHTML = '<div class="stat"><span class="stat-label">暂无数据</span></div>';
                return;
            }}

            var total = data.reduce((sum, item) => sum + item.value, 0);
            var positive = data.filter(item => item.value > 0).reduce((sum, item) => sum + item.value, 0);
            var negative = data.filter(item => item.value < 0).reduce((sum, item) => sum + item.value, 0);

            var html = `
                <div class="stat">
                    <span class="stat-label">总计:</span>
                    <span class="stat-value" style="color: ${{color}}">${{total.toFixed(2)}}</span>
                </div>
                <div class="stat">
                    <span class="stat-label">正向贡献:</span>
                    <span class="stat-value positive">${{positive.toFixed(2)}}</span>
                </div>
                <div class="stat">
                    <span class="stat-label">负向影响:</span>
                    <span class="stat-value negative">${{negative.toFixed(2)}}</span>
                </div>
                <div class="stat">
                    <span class="stat-label">项目数:</span>
                    <span class="stat-value">${{data.length}}</span>
                </div>
            `;
            document.getElementById(elementId).innerHTML = html;
        }}

        // 绘制所有分析图表
        showStats('weibao-stats', analysisData['收入端-维修保养'], 'rgb(0, 114, 109)');
        drawAnalysisChart('weibao-chart', analysisData['收入端-维修保养'], '维修保养', 'rgb(0, 114, 109)');

        showStats('hunhe-stats', analysisData['收入端-混合维修'], 'rgb(206, 164, 114)');
        drawAnalysisChart('hunhe-chart', analysisData['收入端-混合维修'], '混合维修', 'rgb(206, 164, 114)');

        showStats('neibu-stats', analysisData['收入端-内部结算'], 'rgb(0, 114, 109)');
        drawAnalysisChart('neibu-chart', analysisData['收入端-内部结算'], '内部结算', 'rgb(0, 114, 109)');

        showStats('feiyong-stats', analysisData['费用端'], 'rgb(244, 67, 54)');
        drawAnalysisChart('feiyong-chart', analysisData['费用端'], '费用端', 'rgb(244, 67, 54)');

        // 响应式
        window.addEventListener('resize', function() {{
            waterfallChart.resize();
            if (analysisData['收入端-维修保养'].length > 0) {{
                echarts.init(document.getElementById('weibao-chart')).resize();
            }}
            if (analysisData['收入端-混合维修'].length > 0) {{
                echarts.init(document.getElementById('hunhe-chart')).resize();
            }}
            if (analysisData['收入端-内部结算'].length > 0) {{
                echarts.init(document.getElementById('neibu-chart')).resize();
            }}
            if (analysisData['费用端'].length > 0) {{
                echarts.init(document.getElementById('feiyong-chart')).resize();
            }}
        }});
    </script>
</body>
</html>"""

# 保存 HTML 文件
with open(config['output_html_path'], "w", encoding="utf-8") as f
    f.write(html_content)

print(f"Success! Waterfall analysis HTML file generated: {config['output_html_path']}")
print(f"\nData Summary:")
print(f"  Categories: {len(categories)}")
print(f"  Data Range: {min(values):.2f} to {max(values):.2f}")
print(f"\nCategory Analysis:")
print(f"  Revenue-Maintenance: {len(analysis_data['收入端-维修保养'])} items")
print(f"  Revenue-Mixed Repair: {len(analysis_data['收入端-混合维修'])} items")
print(f"  Revenue-Internal Settlement: {len(analysis_data['收入端-内部结算'])} items")
print(f"  Expenses: {len(analysis_data['费用端'])} items")
